import streamlit as st
import pandas as pd
import requests
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side

# ==================================================
# CONFIGURAÇÃO BÁSICA
# ==================================================
st.set_page_config(page_title="Processo de Trocas", layout="wide")
st.title("♻️ Processo de Trocas")

# ==================================================
# ESTADO DA SESSÃO
# ==================================================
if "trocas_dados" not in st.session_state:
    st.session_state.trocas_dados = []

# ==================================================
# VALIDAÇÕES AUXILIARES
# ==================================================
def validar_ean(ean):
    """Valida o EAN antes de qualquer requisição à API."""
    if pd.isna(ean):
        return False, "Código de barras vazio."

    ean = str(ean).strip()

    if not ean.isdigit():
        return False, "O código de barras deve conter apenas números."

    if len(ean) > 14:
        return False, "O código de barras não pode ter mais de 14 dígitos."

    if len(ean) < 1:
        return False, "Código de barras inválido."

    return True, ean.zfill(14)


def validar_quantidade(qtd):
    """Valida a quantidade."""
    try:
        qtd = int(qtd)
        if qtd < 1:
            return False, "Quantidade deve ser pelo menos 1."
        return True, qtd
    except:
        return False, "Quantidade inválida."


# ==================================================
# FUNÇÃO PARA CONSULTA VIA API
# ==================================================
API_HEADERS = {
     "x-api-key": st.secrets["api"]["x_api_key"],
     "Cookie": st.secrets["api"]["cookie"]
}

def buscar_produto_api(ean_input):
    """Consulta no sistema: Produto → Fornecedor → Dados fornecedor com tratamento de erros."""
    try:
        valid, ean_or_msg = validar_ean(ean_input)
        if not valid:
            return None, ean_or_msg

        ean = ean_or_msg

        url_prod = f"https://lojasmimi.varejofacil.com/api/v1/produto/produtos/consulta/{ean}"
        r_prod = requests.get(url_prod, headers=API_HEADERS)

        if r_prod.status_code == 404:
            return None, f"Produto não encontrado (404)."

        produto = r_prod.json()
        produto_id = produto.get("id")
        descricao = produto.get("descricao")

        if not produto_id:
            return None, "Produto não encontrado."

        # fornecedores
        url_forns = f"https://lojasmimi.varejofacil.com/api/v1/produto/produtos/{produto_id}/fornecedores"
        r_forns = requests.get(url_forns, headers=API_HEADERS)
        items = r_forns.json().get("items", [])
        if not items:
            return None, "Nenhum fornecedor encontrado."

        fornecedor_id = items[0].get("fornecedorId")
        referencia = items[0].get("referencia")

        # dados do fornecedor
        url_forn = f"https://lojasmimi.varejofacil.com/api/v1/pessoa/fornecedores/{fornecedor_id}"
        r_forn = requests.get(url_forn, headers=API_HEADERS)
        forn_data = r_forn.json()
        fantasia = forn_data.get("fantasia")

        return {
            "CODIGO BARRA": ean,
            "CODIGO": referencia,
            "DESCRICAO": descricao,
            "FORNECEDOR": fantasia
        }, None

    except requests.exceptions.RequestException:
        return None, "Falha de comunicação com a API. Tente novamente."

    except Exception as e:
        return None, f"Erro inesperado: {e}"


# ==================================================
# FUNÇÃO PARA GERAR FORMULÁRIO EXCEL DINÂMICO
# ==================================================
def gerar_formulario_excel_dinamico(dados, numero_caixa="", responsavel=""):
    """Gera o formulário Excel dinamicamente, sem limite de 27 itens."""
    try:
        modelo_path = "FORM-TROCAS.xlsx"
        wb = load_workbook(modelo_path)
        ws = wb.active
        
        # Definir bordas pretas
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # AJUSTE 1: Adicionar data de hoje na célula E3
        data_hoje = datetime.now().strftime("%d/%m/%Y")
        ws["E3"] = data_hoje
        
        # Verificar quantos itens serão incluídos
        total_itens = len(dados)
        
        # Preencher os produtos dinamicamente a partir da linha 6
        linha_inicial = 6
        
        for i, item in enumerate(dados):
            linha_atual = linha_inicial + i
            
            # Preencher os dados do produto
            ws[f"A{linha_atual}"] = item["CODIGO BARRA"]   # Código de Barras
            ws[f"B{linha_atual}"] = item["CODIGO"]        # Código
            ws[f"C{linha_atual}"] = item["FORNECEDOR"]    # Fornecedor
            ws[f"D{linha_atual}"] = item["DESCRICAO"]     # Descrição
            ws[f"E{linha_atual}"] = item["QUANTIDADE"]    # Quantidade
            
            # Aplicar bordas pretas nas células de A a E
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f"{col}{linha_atual}"].border = thin_border
            
            # AJUSTE: Definir altura da linha para 21.00 (28 pixels)
            ws.row_dimensions[linha_atual].height = 21.00
        
        # Calcular posições dinâmicas para N° CAIXA, RESPONSÁVEL e ASS
        linha_caixa = linha_inicial + total_itens + 1  # +2 da especificação original
        linha_responsavel = linha_inicial + total_itens + 2  # +3 da especificação original
        linha_assinatura = linha_inicial + total_itens + 3  # AJUSTE 2: Nova linha para assinatura
        
        # Preencher N° CAIXA
        ws[f"C{linha_caixa}"] = "N° CAIXA:"
        ws[f"D{linha_caixa}"] = numero_caixa
        
        # Aplicar bordas nas células do CAIXA
        ws[f"C{linha_caixa}"].border = thin_border
        ws[f"D{linha_caixa}"].border = thin_border
        
        # AJUSTE: Definir altura da linha do CAIXA para 21.00
        ws.row_dimensions[linha_caixa].height = 21.00
        
        # Preencher RESPONSÁVEL
        ws[f"C{linha_responsavel}"] = "RESPONSÁVEL:"
        ws[f"D{linha_responsavel}"] = responsavel
        
        # Aplicar bordas nas células do RESPONSÁVEL
        ws[f"C{linha_responsavel}"].border = thin_border
        ws[f"D{linha_responsavel}"].border = thin_border
        
        # AJUSTE: Definir altura da linha do RESPONSÁVEL para 21.00
        ws.row_dimensions[linha_responsavel].height = 21.00
        
        # AJUSTE 2: Preencher linha de ASSINATURA
        ws[f"C{linha_assinatura}"] = "ASS:"
        # A célula D da linha de assinatura fica vazia (para assinatura)
        
        # Aplicar bordas nas células da ASSINATURA
        ws[f"C{linha_assinatura}"].border = thin_border
        ws[f"D{linha_assinatura}"].border = thin_border
        
        # AJUSTE: Definir altura da linha da ASSINATURA para 21.00
        ws.row_dimensions[linha_assinatura].height = 21.00

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Retorna o arquivo
        return output, total_itens, None

    except Exception as e:
        return None, 0, f"Erro ao gerar formulário: {str(e)}"


# ==================================================
# FUNÇÃO PARA GERAR ARQUIVO MODELO DO LOTE
# ==================================================
def gerar_modelo_lote():
    wb = Workbook()
    ws = wb.active
    ws.title = "TROCAS"

    ws["A1"] = "CODIGO DE BARRAS"
    ws["B1"] = "QUANTIDADE"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ==================================================
# 🟦 ABAS DO SISTEMA
# ==================================================
tab1, tab2, tab3 = st.tabs(["🔍 INDIVIDUAL", "📦 LOTE", "📋 RELATÓRIO"])

# ==================================================
# 1️⃣ INDIVIDUAL
# ==================================================
with tab1:
    st.subheader("🔍 Buscar Produto Para Troca")

    col1, col2 = st.columns([4, 2])
    ean_input = col1.text_input("Digite o Código de Barras (EAN):")
    quantidade = col2.number_input("Quantidade", min_value=1, step=1, value=1)

    if st.button("🔎 Buscar Produto"):
        # Primeiro validar a quantidade ANTES de buscar o produto
        valid_qtd, qtd_or_msg = validar_quantidade(quantidade)
        if not valid_qtd:
            st.error(qtd_or_msg)
            # Se quantidade inválida, para aqui e NÃO busca o produto
            st.stop()
        
        # Agora buscar o produto
        resultado, erro = buscar_produto_api(ean_input)
        if erro:
            st.error(erro)
        else:
            # evitar duplicados
            if any(p["CODIGO BARRA"] == resultado["CODIGO BARRA"] for p in st.session_state.trocas_dados):
                st.warning("⚠️ Produto já estava na lista. Quantidade somada.")
                for p in st.session_state.trocas_dados:
                    if p["CODIGO BARRA"] == resultado["CODIGO BARRA"]:
                        p["QUANTIDADE"] += qtd_or_msg
            else:
                resultado["QUANTIDADE"] = qtd_or_msg
                st.session_state.trocas_dados.append(resultado)

            st.success(f"✅ Produto adicionado: {resultado['DESCRICAO']}")


# ==================================================
# 2️⃣ LOTE
# ==================================================
with tab2:
    st.subheader("📦 Lançar Trocas em Lote")

    st.markdown("### 📤 Baixar modelo Excel")
    st.download_button(
        label="📥 Baixar Modelo Excel",
        data=gerar_modelo_lote(),
        file_name="MODELO_TROCAS_LOTE.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.markdown("---")

    uploaded_file = st.file_uploader("📁 Envie o arquivo preenchido", type=["xlsx"])

    if uploaded_file:
        try:
            df_lote = pd.read_excel(uploaded_file)
        except:
            st.error("❌ Arquivo corrompido ou ilegível.")
            st.stop()

        if "CODIGO DE BARRAS" not in df_lote or "QUANTIDADE" not in df_lote:
            st.error("❌ O arquivo deve conter 'CODIGO DE BARRAS' e 'QUANTIDADE'.")
            st.stop()

        st.success("Arquivo carregado!")

        if df_lote["CODIGO DE BARRAS"].duplicated().any():
            st.warning("⚠️ Códigos duplicados encontrados — as quantidades serão somados.")
            df_lote = df_lote.groupby("CODIGO DE BARRAS", as_index=False)["QUANTIDADE"].sum()

        if st.button("🚀 Processar Lote"):
            sucessos = []
            falhas = []
            fornecedores = set()

            progress = st.progress(0)
            total = len(df_lote)

            for i, row in df_lote.iterrows():

                valid_ean, ean_or_msg = validar_ean(row["CODIGO DE BARRAS"])
                if not valid_ean:
                    falhas.append({"CODIGO": row["CODIGO DE BARRAS"], "ERRO": ean_or_msg})
                    progress.progress((i + 1) / total)
                    continue

                valid_qtd, qtd_or_msg = validar_quantidade(row["QUANTIDADE"])
                if not valid_qtd:
                    falhas.append({"CODIGO": row["CODIGO DE BARRAS"], "ERRO": qtd_or_msg})
                    progress.progress((i + 1) / total)
                    continue

                resultado, erro = buscar_produto_api(ean_or_msg)

                if erro:
                    falhas.append({"CODIGO": row["CODIGO DE BARRAS"], "ERRO": erro})
                else:
                    fornecedores.add(resultado["FORNECEDOR"])
                    resultado["QUANTIDADE"] = qtd_or_msg
                    sucessos.append(resultado)

                progress.progress((i + 1) / total)

            # REMOVIDA VALIDAÇÃO DE MÚLTIPLOS FORNECEDORES - Permitido agora
            
            st.subheader("📊 Resultado do Lote")
            st.success(f"✅ Sucessos: {len(sucessos)}")
            st.error(f"❌ Falhas: {len(falhas)}")

            if falhas:
                st.write("### ❌ Erros encontrados")
                st.dataframe(pd.DataFrame(falhas))

            # adicionar ao relatório
            for item in sucessos:
                if any(p["CODIGO BARRA"] == item["CODIGO BARRA"] for p in st.session_state.trocas_dados):
                    for p in st.session_state.trocas_dados:
                        if p["CODIGO BARRA"] == item["CODIGO BARRA"]:
                            p["QUANTIDADE"] += item["QUANTIDADE"]
                else:
                    st.session_state.trocas_dados.append(item)

            st.success("🎉 Produtos válidos adicionados ao relatório!")


# ==================================================
# 3️⃣ RELATÓRIO
# ==================================================
with tab3:
    st.subheader("📋 Produtos Adicionados Para Troca")

    if st.session_state.trocas_dados:
        df_trocas = pd.DataFrame(st.session_state.trocas_dados)
        
        # Verificar quantos fornecedores diferentes existem
        fornecedores = df_trocas["FORNECEDOR"].unique()
        if len(fornecedores) > 1:
            st.info(f"📦 **Múltiplos fornecedores detectados:** {len(fornecedores)} fornecedores diferentes")
        
        st.dataframe(df_trocas, width="stretch")

        # NOVO: Campos para N° CAIXA e RESPONSÁVEL
        st.write("### 📝 Informações Adicionais do Formulário")
        
        col1, col2 = st.columns(2)
        
        with col1:
            numero_caixa = st.text_input("N° CAIXA:", placeholder="Digite o número da caixa")
        
        with col2:
            responsavel = st.text_input("RESPONSÁVEL:", placeholder="Digite o nome do responsável")

        colA, colB = st.columns([1, 3])

        if colA.button("🗑️ Remover Último Item"):
            removido = st.session_state.trocas_dados.pop()
            st.warning(f"Item removido: {removido['DESCRICAO']} (Qtd: {removido['QUANTIDADE']})")

        if colB.button("📄 Gerar Formulário de Troca"):
            
            # Gerar o formulário DINÂMICO
            excel_bytes, total_processado, erro = gerar_formulario_excel_dinamico(
                st.session_state.trocas_dados, 
                numero_caixa, 
                responsavel
            )

            if erro:
                st.error(erro)
            else:
                # APENAS A MENSAGEM SIMPLES DE SUCESSO
                st.success("✅ Formulário gerado com sucesso")
                
                # Botão de download (mantido como estava)
                st.download_button(
                    label="📥 Baixar Formulário Dinâmico",
                    data=excel_bytes,
                    file_name=f"FORMULARIO_TROCAS_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.info("Nenhum produto adicionado ainda.")

# ==================================================
# RODAPÉ
# ==================================================
st.markdown("""
<hr style='border: 0; height: 1px; background: #ccc; margin-top: 2em; margin-bottom: 1em;' />
<div style='text-align: center; color: grey; font-size: 0.8em;'>
    Aplicativo desenvolvido por <a href="https://github.com/opablodantas" target="_blank"><strong>PABLO</strong></a> para as lojas <strong>MIMI</strong>. Todos os direitos reservados.
</div>
""", unsafe_allow_html=True)